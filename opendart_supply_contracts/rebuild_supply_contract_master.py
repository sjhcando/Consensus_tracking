from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook

from build_contracts_for_stocks import OUTDIR, style_sheet
from daily_update_supply_contracts import MASTER_FILE, MASTER_HEADERS


def read_meta(wb):
    if len(wb.worksheets) < 3:
        return "", ""
    ws = wb.worksheets[2]
    meta = {}
    for r in range(1, ws.max_row + 1):
        meta[str(ws.cell(r, 1).value or "")] = ws.cell(r, 2).value
    return str(meta.get("회사") or ""), str(meta.get("종목코드") or "")


def collect_company_file(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    company, ticker = read_meta(wb)
    if not company:
        company = path.stem
    rows = defaultdict(lambda: {"new": 0, "corr": 0, "receipts": []})

    if len(wb.worksheets) >= 1:
        ws = wb.worksheets[0]
        for r in range(2, ws.max_row + 1):
            date = ws.cell(r, 1).value
            rcp = ws.cell(r, 8).value
            if date and rcp:
                rows[str(date)]["new"] += 1
                rows[str(date)]["receipts"].append(str(rcp))

    if len(wb.worksheets) >= 2:
        ws = wb.worksheets[1]
        seen_corr_receipts = defaultdict(set)
        for r in range(2, ws.max_row + 1):
            date = ws.cell(r, 1).value
            rcp = ws.cell(r, 7).value
            if date and rcp:
                rows[str(date)]["corr"] += 1
                seen_corr_receipts[str(date)].add(str(rcp))
        for date, receipts in seen_corr_receipts.items():
            rows[date]["receipts"].extend(sorted(receipts))

    return company, ticker, rows


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "일자별 공급계약"
    ws.append(MASTER_HEADERS)

    for path in sorted(OUTDIR.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        if path.name == MASTER_FILE.name:
            continue
        company, ticker, rows = collect_company_file(path)
        for date in sorted(rows):
            data = rows[date]
            if data["new"] or data["corr"]:
                ws.append([
                    date,
                    company,
                    ticker,
                    data["new"],
                    data["corr"],
                    ", ".join(dict.fromkeys(data["receipts"])),
                    str(path),
                ])

    style_sheet(ws)
    OUTDIR.mkdir(parents=True, exist_ok=True)
    wb.save(MASTER_FILE)
    print(MASTER_FILE)
    print(f"rows={ws.max_row - 1}")


if __name__ == "__main__":
    main()
