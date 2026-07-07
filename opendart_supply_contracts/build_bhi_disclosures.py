import json
import re
import zipfile
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
DOCS = ROOT / "docs"
OUTDIR = ROOT / "outputs" / "bhi_opendart"
OUTFILE = OUTDIR / "bhi_single_sales_contracts_2023_to_20260701.xlsx"


class TableParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.tables = []
        self._table_stack = []
        self._row = None
        self._cell = None
        self._in_span = False

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._table_stack.append([])
        elif tag == "tr" and self._table_stack:
            self._row = []
        elif tag == "td" and self._row is not None:
            self._cell = []
        elif tag == "br" and self._cell is not None:
            self._cell.append("\n")

    def handle_endtag(self, tag):
        if tag == "td" and self._cell is not None and self._row is not None:
            text = re.sub(r"[ \t\r\f\v]+", " ", "".join(self._cell))
            text = re.sub(r" *\n *", "\n", text).strip()
            self._row.append(text)
            self._cell = None
        elif tag == "tr" and self._row is not None and self._table_stack:
            if any(c for c in self._row):
                self._table_stack[-1].append(self._row)
            self._row = None
        elif tag == "table" and self._table_stack:
            table = self._table_stack.pop()
            if table:
                if self._table_stack:
                    self._table_stack[-1].extend(table)
                else:
                    self.tables.append(table)

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)


def read_doc_html(rcept_no):
    zpath = DOCS / f"{rcept_no}.zip"
    with zipfile.ZipFile(zpath) as zf:
        name = next(n for n in zf.namelist() if n.lower().endswith(".xml"))
        data = zf.read(name)
    for enc in ("euc-kr", "utf-8", "cp949"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            pass
    return data.decode("utf-8", errors="replace")


def parse_tables(html):
    p = TableParser()
    p.feed(html)
    return p.tables


def clean_label(s):
    return re.sub(r"^\d+\.\s*", "", s or "").strip()


def find_value(rows, label, sublabel=None):
    for row in rows:
        cells = [clean_label(c) for c in row]
        if sublabel:
            if len(cells) >= 2 and sublabel in cells[-2]:
                return row[-1].strip()
            if any(sublabel in c for c in cells[:-1]):
                return row[-1].strip()
            if any(label in c for c in cells) and any(sublabel in c for c in cells):
                return row[-1].strip() if row else ""
        else:
            for i, c in enumerate(cells):
                if label in c:
                    if len(row) > i + 1:
                        return row[-1].strip()
    return ""


def parse_new_fields(tables):
    body = max(tables, key=lambda t: sum(len(r) for r in t))
    rows = body
    amount_candidates = [
        find_value(rows, "계약금액 총액"),
        find_value(rows, "계약금액(원)"),
        find_value(rows, "확정 계약금액"),
        find_value(rows, "계약금액"),
    ]
    amount = next((v for v in amount_candidates if to_int_amount(v) != v), "")
    if not amount:
        amount = next((v for v in amount_candidates if v), "")
    return {
        "계약상대방": find_value(rows, "계약상대방") or find_value(rows, "계약상대"),
        "판매공급지역": find_value(rows, "판매ㆍ공급지역"),
        "계약금액": amount,
        "계약시작일": find_value(rows, "계약기간", "시작일") or find_value(rows, "시작일"),
        "계약종료일": find_value(rows, "계약기간", "종료일") or find_value(rows, "종료일"),
    }


def parse_correction_fields(tables):
    corr_table = None
    for table in tables:
        flat = "\n".join(" | ".join(r) for r in table)
        if "정정관련 공시서류제출일" in flat and "정정전" in flat and "정정후" in flat:
            corr_table = table
            break
    if not corr_table:
        return []
    related_date = find_value(corr_table, "정정관련 공시서류제출일")
    rows = []
    seen_header = False
    for row in corr_table:
        cells = [c.strip() for c in row]
        if len(cells) >= 3 and "정정항목" in cells[0] and "정정전" in cells[1] and "정정후" in cells[2]:
            seen_header = True
            continue
        if seen_header and len(cells) >= 3:
            item, before, after = cells[0], cells[1], cells[2]
            if item and item not in ("-", "4. 정정사항"):
                rows.append({
                    "정정관련 공시서류제출일": related_date,
                    "정정항목": item,
                    "정정전": before,
                    "정정후": after,
                })
    if not rows:
        rows.append({
            "정정관련 공시서류제출일": related_date,
            "정정항목": "",
            "정정전": "",
            "정정후": "",
        })
    return rows


def fmt_date(s):
    if not s:
        return ""
    s = str(s)
    if re.fullmatch(r"\d{8}", s):
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    return s


def to_int_amount(s):
    text = (s or "").replace(",", "").strip()
    return int(text) if re.fullmatch(r"-?\d+", text) else s


def main():
    items = json.loads((ROOT / "filtered_list.json").read_text(encoding="utf-8-sig"))
    new_rows, corr_rows, errors = [], [], []
    for it in items:
        rcp = it["rcept_no"]
        try:
            html = read_doc_html(rcp)
            tables = parse_tables(html)
            source = f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcp}"
            is_corr = "기재정정" in it["report_nm"] or "첨부추가" in it["report_nm"]
            if is_corr:
                for row in parse_correction_fields(tables):
                    corr_rows.append({
                        "공시일": fmt_date(it["rcept_dt"]),
                        **row,
                        "공시명": it["report_nm"].strip(),
                        "접수번호": rcp,
                        "출처URL": source,
                    })
            else:
                fields = parse_new_fields(tables)
                fields["계약금액"] = to_int_amount(fields["계약금액"])
                new_rows.append({
                    "공시일": fmt_date(it["rcept_dt"]),
                    **fields,
                    "공시명": it["report_nm"].strip(),
                    "접수번호": rcp,
                    "출처URL": source,
                })
        except Exception as e:
            errors.append({"접수번호": rcp, "공시일": fmt_date(it["rcept_dt"]), "공시명": it["report_nm"], "오류": str(e)})

    OUTDIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "신규 공시"
    write_sheet(ws, new_rows, ["공시일", "계약상대방", "판매공급지역", "계약금액", "계약시작일", "계약종료일", "공시명", "접수번호", "출처URL"])
    ws2 = wb.create_sheet("정정 공시")
    write_sheet(ws2, corr_rows, ["공시일", "정정관련 공시서류제출일", "정정항목", "정정전", "정정후", "공시명", "접수번호", "출처URL"])
    ws3 = wb.create_sheet("수집 기준")
    meta = [
        ("회사", "비에이치아이"),
        ("종목코드", "083650"),
        ("OpenDART 고유번호", "00409788"),
        ("수집기간", "2023-01-01 ~ 2026-07-01"),
        ("대상 공시명", "단일판매ㆍ공급계약체결 및 정정/첨부추가"),
        ("신규 공시 수", len(new_rows)),
        ("정정 공시 행 수", len(corr_rows)),
        ("원문 파싱 오류 수", len(errors)),
    ]
    for r, (k, v) in enumerate(meta, 1):
        ws3.cell(r, 1, k)
        ws3.cell(r, 2, v)
    if errors:
        ws4 = wb.create_sheet("파싱 오류")
        write_sheet(ws4, errors, ["접수번호", "공시일", "공시명", "오류"])
    for sheet in wb.worksheets:
        style_sheet(sheet)
    wb.save(OUTFILE)
    print(OUTFILE)
    print(f"new={len(new_rows)} correction_rows={len(corr_rows)} errors={len(errors)}")


def write_sheet(ws, rows, headers):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), max((len(x) for x in value.split("\n")), default=0) + 2), 60)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = max(width, 10)
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 34


if __name__ == "__main__":
    main()
