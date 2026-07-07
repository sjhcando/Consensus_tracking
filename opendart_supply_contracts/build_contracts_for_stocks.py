import json
import os
import re
import time
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from build_bhi_disclosures import (
    parse_correction_fields,
    parse_new_fields,
    parse_tables,
    to_int_amount,
)


ROOT = Path(__file__).resolve().parent
STOCKS_JSON = Path(r"C:\투자\stocks.json")
CORP_XML = ROOT / "corpCode" / "CORPCODE.xml"
DOCS = ROOT / "docs_batch"
OUTDIR = Path(r"G:\내 드라이브\3. Stocks\Open Dart\공급계약")
API_KEY = os.environ.get("OPENDART_API_KEY", "")
BEGIN_DATE = "20230101"
END_DATE = "20260701"


def fmt_date(s):
    s = str(s or "")
    return f"{s[:4]}-{s[4:6]}-{s[6:]}" if re.fullmatch(r"\d{8}", s) else s


def safe_filename(name):
    cleaned = re.sub(r'[\\/:*?"<>|]', "_", name).strip()
    return cleaned or "unknown"


def read_json_url(url, retries=3):
    ensure_api_key()
    last = None
    for i in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=30) as res:
                return json.loads(res.read().decode("utf-8"))
        except Exception as exc:
            last = exc
            time.sleep(1.5 * (i + 1))
    raise last


def download_url(url, path, retries=3):
    if path.exists() and path.stat().st_size > 500:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    last = None
    for i in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=45) as res:
                path.write_bytes(res.read())
            if path.stat().st_size > 500:
                return
        except Exception as exc:
            last = exc
            time.sleep(1.5 * (i + 1))
    raise last or RuntimeError(f"download failed: {url}")


def load_stocks():
    data = json.loads(STOCKS_JSON.read_text(encoding="utf-8-sig"))
    rows = []
    for sector, items in data.items():
        for item in items:
            rows.append({"sector": sector, "name": item["name"], "ticker": item["ticker"]})
    return rows


def ensure_api_key():
    if not API_KEY:
        raise RuntimeError("OPENDART_API_KEY environment variable is required")


def ensure_corp_code_file():
    if CORP_XML.exists():
        return
    ensure_api_key()
    CORP_XML.parent.mkdir(parents=True, exist_ok=True)
    url = "https://opendart.fss.or.kr/api/corpCode.xml?" + urllib.parse.urlencode({"crtfc_key": API_KEY})
    zip_path = CORP_XML.parent / "corpCode.zip"
    download_url(url, zip_path)
    with zipfile.ZipFile(zip_path) as zf:
        zf.extractall(CORP_XML.parent)


def load_corp_codes():
    ensure_corp_code_file()
    tree = ET.parse(CORP_XML)
    mapping = {}
    for node in tree.getroot().findall("list"):
        stock = (node.findtext("stock_code") or "").strip()
        if stock:
            mapping[stock] = {
                "corp_code": (node.findtext("corp_code") or "").strip(),
                "corp_name": (node.findtext("corp_name") or "").strip(),
            }
    return mapping


def fetch_disclosure_list(corp_code, begin_date=None, end_date=None):
    all_items = []
    page = 1
    begin_date = begin_date or BEGIN_DATE
    end_date = end_date or END_DATE
    while True:
        params = {
            "crtfc_key": API_KEY,
            "corp_code": corp_code,
            "bgn_de": begin_date,
            "end_de": end_date,
            "page_no": page,
            "page_count": 100,
        }
        url = "https://opendart.fss.or.kr/api/list.json?" + urllib.parse.urlencode(params)
        data = read_json_url(url)
        if data.get("status") not in ("000", "013"):
            raise RuntimeError(f"{data.get('status')} {data.get('message')}")
        all_items.extend(data.get("list") or [])
        total_page = int(data.get("total_page") or 1)
        if page >= total_page:
            break
        page += 1
        time.sleep(0.25)
    return [
        item
        for item in all_items
        if "단일판매ㆍ공급계약체결" in (item.get("report_nm") or "")
    ]


def read_doc_html(rcept_no):
    zpath = DOCS / f"{rcept_no}.zip"
    with zipfile.ZipFile(zpath) as zf:
        name = next(n for n in zf.namelist() if n.lower().endswith(".xml"))
        raw = zf.read(name)
    for enc in ("euc-kr", "utf-8", "cp949"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def collect_company(stock):
    corp_code = stock["corp_code"]
    items = fetch_disclosure_list(corp_code)
    new_rows, corr_rows, errors = [], [], []
    for item in items:
        rcp = item["rcept_no"]
        try:
            doc_url = (
                "https://opendart.fss.or.kr/api/document.xml?"
                + urllib.parse.urlencode({"crtfc_key": API_KEY, "rcept_no": rcp})
            )
            download_url(doc_url, DOCS / f"{rcp}.zip")
            html = read_doc_html(rcp)
            tables = parse_tables(html)
            source = f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcp}"
            report_nm = (item.get("report_nm") or "").strip()
            is_corr = "기재정정" in report_nm or "첨부추가" in report_nm
            if is_corr:
                for row in parse_correction_fields(tables):
                    corr_rows.append({
                        "공시일": fmt_date(item["rcept_dt"]),
                        **row,
                        "공시명": report_nm,
                        "접수번호": rcp,
                        "출처URL": source,
                    })
            else:
                fields = parse_new_fields(tables)
                fields["계약금액"] = to_int_amount(fields["계약금액"])
                if fields["계약금액"] == "-":
                    fields["계약금액"] = ""
                new_rows.append({
                    "공시일": fmt_date(item["rcept_dt"]),
                    **fields,
                    "공시명": report_nm,
                    "접수번호": rcp,
                    "출처URL": source,
                })
            time.sleep(0.15)
        except Exception as exc:
            errors.append({
                "접수번호": rcp,
                "공시일": fmt_date(item.get("rcept_dt")),
                "공시명": item.get("report_nm"),
                "오류": str(exc),
            })
    return items, new_rows, corr_rows, errors


def write_sheet(ws, rows, headers):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            max_line = max((len(x) for x in value.split("\n")), default=0)
            widths[cell.column] = min(max(widths.get(cell.column, 0), max_line + 2), 60)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = max(width, 10)
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 34


def save_workbook(stock, items, new_rows, corr_rows, errors):
    wb = Workbook()
    ws = wb.active
    ws.title = "신규 공시"
    write_sheet(ws, new_rows, ["공시일", "계약상대방", "판매공급지역", "계약금액", "계약시작일", "계약종료일", "공시명", "접수번호", "출처URL"])
    ws2 = wb.create_sheet("정정 공시")
    write_sheet(ws2, corr_rows, ["공시일", "정정관련 공시서류제출일", "정정항목", "정정전", "정정후", "공시명", "접수번호", "출처URL"])
    ws3 = wb.create_sheet("수집 기준")
    meta = [
        ("회사", stock["name"]),
        ("종목코드", stock["ticker"]),
        ("OpenDART 고유번호", stock.get("corp_code", "")),
        ("수집기간", f"{fmt_date(BEGIN_DATE)} ~ {fmt_date(END_DATE)}"),
        ("대상 공시명", "단일판매ㆍ공급계약체결 및 정정/첨부추가"),
        ("대상 공시 수", len(items)),
        ("신규 공시 수", len(new_rows)),
        ("정정 공시 행 수", len(corr_rows)),
        ("원문 파싱 오류 수", len(errors)),
    ]
    for r, (k, v) in enumerate(meta, 1):
        ws3.cell(r, 1, k)
        ws3.cell(r, 2, v)
    if errors:
        ws4 = wb.create_sheet("파싱 오류")
        write_sheet(ws4, errors, ["접수번호", "공시일", "공시명", "오류"])
    for sheet in wb.worksheets:
        style_sheet(sheet)
    OUTDIR.mkdir(parents=True, exist_ok=True)
    path = OUTDIR / f"{safe_filename(stock['name'])}.xlsx"
    wb.save(path)
    return path


def main():
    stocks = load_stocks()
    corp_codes = load_corp_codes()
    summary = []
    for stock in stocks:
        info = corp_codes.get(stock["ticker"])
        if not info:
            summary.append({**stock, "file": "", "new": 0, "corr": 0, "errors": 1, "note": "corp_code not found"})
            continue
        stock = {**stock, **info}
        items, new_rows, corr_rows, errors = collect_company(stock)
        path = save_workbook(stock, items, new_rows, corr_rows, errors)
        summary.append({
            "sector": stock["sector"],
            "name": stock["name"],
            "ticker": stock["ticker"],
            "file": str(path),
            "new": len(new_rows),
            "corr": len(corr_rows),
            "errors": len(errors),
        })
        print(f"{stock['name']} {stock['ticker']}: 신규 {len(new_rows)}, 정정 {len(corr_rows)}, 오류 {len(errors)}")
        time.sleep(0.35)
    (OUTDIR / "_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"saved {len(summary)} company files to {OUTDIR}")


if __name__ == "__main__":
    main()
