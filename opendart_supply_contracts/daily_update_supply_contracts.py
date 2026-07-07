import argparse
import json
import re
import time
import urllib.parse
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

from build_bhi_disclosures import parse_correction_fields, parse_new_fields, parse_tables, to_int_amount
from build_contracts_for_stocks import (
    API_KEY,
    DOCS,
    OUTDIR,
    download_url,
    fetch_disclosure_list,
    fmt_date,
    load_corp_codes,
    load_stocks,
    read_doc_html,
    safe_filename,
    style_sheet,
    write_sheet,
)


ROOT = Path(__file__).resolve().parent
MASTER_FILE = OUTDIR / "공급계약_마스터.xlsx"
LOG_DIR = ROOT / "logs"

NEW_HEADERS = ["공시일", "계약상대방", "판매공급지역", "계약금액", "계약시작일", "계약종료일", "공시명", "접수번호", "출처URL"]
CORR_HEADERS = ["공시일", "정정관련 공시서류제출일", "정정항목", "정정전", "정정후", "공시명", "접수번호", "출처URL"]
MASTER_HEADERS = ["공시일", "기업명", "종목코드", "신규공시건수", "정정공시행수", "접수번호", "기업파일"]


def yyyymmdd(value):
    return re.sub(r"\D", "", value)


def today_yyyymmdd():
    return datetime.now().strftime("%Y%m%d")


def is_weekday(date_text):
    dt = datetime.strptime(date_text, "%Y%m%d")
    return dt.weekday() < 5


def lookback_start_date(date_text, business_days=1):
    dt = datetime.strptime(date_text, "%Y%m%d")
    remaining = business_days
    cur = dt
    while remaining > 0:
        cur -= timedelta(days=1)
        if cur.weekday() < 5:
            remaining -= 1
    return cur.strftime("%Y%m%d")


def ensure_company_workbook(path, stock):
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "신규 공시"
    write_sheet(ws, [], NEW_HEADERS)
    ws2 = wb.create_sheet("정정 공시")
    write_sheet(ws2, [], CORR_HEADERS)
    ws3 = wb.create_sheet("수집 기준")
    meta = [
        ("회사", stock["name"]),
        ("종목코드", stock["ticker"]),
        ("OpenDART 고유번호", stock.get("corp_code", "")),
        ("관리방식", "일별 누적 업데이트"),
    ]
    for r, (k, v) in enumerate(meta, 1):
        ws3.cell(r, 1, k)
        ws3.cell(r, 2, v)
    return wb


def existing_receipts(ws):
    if ws.max_row < 2:
        return set()
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "접수번호" not in headers:
        return set()
    col = headers.index("접수번호") + 1
    return {str(ws.cell(r, col).value) for r in range(2, ws.max_row + 1) if ws.cell(r, col).value}


def append_row(ws, headers, row):
    ws.append([row.get(h, "") for h in headers])
    amount_col = headers.index("계약금액") + 1 if "계약금액" in headers else None
    if amount_col:
        cell = ws.cell(ws.max_row, amount_col)
        if cell.value not in (None, ""):
            cell.number_format = "0"


def parse_disclosure(item):
    rcp = item["rcept_no"]
    doc_url = "https://opendart.fss.or.kr/api/document.xml?" + urllib.parse.urlencode({"crtfc_key": API_KEY, "rcept_no": rcp})
    download_url(doc_url, DOCS / f"{rcp}.zip")
    html = read_doc_html(rcp)
    tables = parse_tables(html)
    report_nm = (item.get("report_nm") or "").strip()
    source = f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcp}"
    is_corr = "기재정정" in report_nm or "첨부추가" in report_nm
    if is_corr:
        rows = []
        for row in parse_correction_fields(tables):
            rows.append({
                "공시일": fmt_date(item["rcept_dt"]),
                **row,
                "공시명": report_nm,
                "접수번호": rcp,
                "출처URL": source,
            })
        return [], rows

    fields = parse_new_fields(tables)
    fields["계약금액"] = to_int_amount(fields["계약금액"])
    if fields["계약금액"] == "-":
        fields["계약금액"] = ""
    return [{
        "공시일": fmt_date(item["rcept_dt"]),
        **fields,
        "공시명": report_nm,
        "접수번호": rcp,
        "출처URL": source,
    }], []


def load_master():
    if MASTER_FILE.exists():
        return load_workbook(MASTER_FILE)
    wb = Workbook()
    ws = wb.active
    ws.title = "일자별 공급계약"
    write_sheet(ws, [], MASTER_HEADERS)
    return wb


def append_master(date_text, stock, new_count, corr_count, receipts, company_file):
    wb = load_master()
    ws = wb.worksheets[0]
    existing = set()
    for r in range(2, ws.max_row + 1):
        existing.add((str(ws.cell(r, 1).value), str(ws.cell(r, 3).value)))
    key = (fmt_date(date_text), stock["ticker"])
    if key not in existing:
        ws.append([
            fmt_date(date_text),
            stock["name"],
            stock["ticker"],
            new_count,
            corr_count,
            ", ".join(receipts),
            str(company_file),
        ])
    else:
        for r in range(2, ws.max_row + 1):
            if (str(ws.cell(r, 1).value), str(ws.cell(r, 3).value)) == key:
                ws.cell(r, 4).value = int(ws.cell(r, 4).value or 0) + new_count
                ws.cell(r, 5).value = int(ws.cell(r, 5).value or 0) + corr_count
                old = str(ws.cell(r, 6).value or "")
                merged = [x for x in old.split(", ") if x] + receipts
                ws.cell(r, 6).value = ", ".join(dict.fromkeys(merged))
                break
    style_sheet(ws)
    OUTDIR.mkdir(parents=True, exist_ok=True)
    wb.save(MASTER_FILE)


def update_for_date(date_text, force_weekend=False, lookback_business_days=1):
    date_text = yyyymmdd(date_text)
    if not re.fullmatch(r"\d{8}", date_text):
        raise ValueError("date must be YYYYMMDD or YYYY-MM-DD")
    if not force_weekend and not is_weekday(date_text):
        return []
    start_date = lookback_start_date(date_text, lookback_business_days)

    corp_codes = load_corp_codes()
    results = []
    for stock in load_stocks():
        info = corp_codes.get(stock["ticker"])
        if not info:
            continue
        stock = {**stock, **info}
        items = fetch_disclosure_list(stock["corp_code"], start_date, date_text)
        if not items:
            continue

        company_file = OUTDIR / f"{safe_filename(stock['name'])}.xlsx"
        wb = ensure_company_workbook(company_file, stock)
        new_ws = wb["신규 공시"]
        corr_ws = wb["정정 공시"]
        seen = existing_receipts(new_ws) | existing_receipts(corr_ws)
        added_new = 0
        added_corr = 0
        receipts = []
        new_details = []
        grouped = {}

        for item in sorted(items, key=lambda x: (x.get("rcept_dt", ""), x.get("rcept_no", ""))):
            rcp = item["rcept_no"]
            if rcp in seen:
                continue
            new_rows, corr_rows = parse_disclosure(item)
            disclosure_date = item.get("rcept_dt") or date_text
            bucket = grouped.setdefault(disclosure_date, {
                "new": 0,
                "correction": 0,
                "new_details": [],
                "receipts": [],
            })
            for row in new_rows:
                append_row(new_ws, NEW_HEADERS, row)
                added_new += 1
                detail = {
                    "counterparty": row.get(NEW_HEADERS[1], ""),
                    "region": row.get(NEW_HEADERS[2], ""),
                    "amount": row.get(NEW_HEADERS[3], ""),
                    "start_date": row.get(NEW_HEADERS[4], ""),
                    "end_date": row.get(NEW_HEADERS[5], ""),
                    "receipt": row.get(NEW_HEADERS[7], rcp),
                }
                new_details.append(detail)
                bucket["new"] += 1
                bucket["new_details"].append(detail)
            for row in corr_rows:
                append_row(corr_ws, CORR_HEADERS, row)
                added_corr += 1
                bucket["correction"] += 1
            receipts.append(rcp)
            bucket["receipts"].append(rcp)
            seen.add(rcp)
            time.sleep(0.15)

        if added_new or added_corr:
            for ws in wb.worksheets:
                style_sheet(ws)
            OUTDIR.mkdir(parents=True, exist_ok=True)
            wb.save(company_file)
            for disclosure_date, data in sorted(grouped.items()):
                append_master(disclosure_date, stock, data["new"], data["correction"], data["receipts"], company_file)
                results.append({
                    "date": fmt_date(disclosure_date),
                    "company": stock["name"],
                    "ticker": stock["ticker"],
                    "new": data["new"],
                    "correction": data["correction"],
                    "new_details": data["new_details"],
                    "receipts": data["receipts"],
                    "file": str(company_file),
                })
    return results


def format_amount(value):
    if value in (None, ""):
        return "\uc6d0\ubb38 \uae08\uc561 \ube44\uacf5\uac1c"
    if isinstance(value, (int, float)):
        return f"{int(value):,}\uc6d0"
    return str(value)


def format_briefing(date_text, results):
    display_date = fmt_date(yyyymmdd(date_text))
    if not results:
        return f"{display_date} \uacf5\uae09\uacc4\uc57d \uacf5\uc2dc \ud655\uc778 \uacb0\uacfc: \ucd94\uac00 \ubc18\uc601\ub41c \uc2e0\uaddc/\uc815\uc815 \uacf5\uc2dc \uc5c6\uc74c."

    lines = [f"{display_date} \uacf5\uae09\uacc4\uc57d \uacf5\uc2dc \ud655\uc778 \uacb0\uacfc:"]
    for result in results:
        company = result["company"]
        result_date = result.get("date", display_date)
        if result.get("new"):
            details = result.get("new_details") or []
            if details:
                for detail in details:
                    lines.append(
                        f"- {result_date} {company}: \uc2e0\uaddc \uacc4\uc57d \uacf5\uc2dc 1\uac74, "
                        f"\uacc4\uc57d\uc0c1\ub300\ubc29 {detail.get('counterparty') or '-'}, "
                        f"\uacc4\uc57d\uae08\uc561 {format_amount(detail.get('amount'))}"
                    )
            else:
                lines.append(f"- {result_date} {company}: \uc2e0\uaddc \uacc4\uc57d \uacf5\uc2dc {result['new']}\uac74")
        if result.get("correction"):
            lines.append(f"- {result_date} {company}: \uc815\uc815 \uacf5\uc2dc {result['correction']}\ud589")
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=today_yyyymmdd(), help="YYYYMMDD or YYYY-MM-DD. Default: today.")
    parser.add_argument("--force-weekend", action="store_true")
    parser.add_argument("--lookback-business-days", type=int, default=1)
    args = parser.parse_args()

    LOG_DIR.mkdir(exist_ok=True)
    results = update_for_date(args.date, args.force_weekend, args.lookback_business_days)
    log_path = LOG_DIR / f"daily_update_{yyyymmdd(args.date)}.json"
    briefing_path = LOG_DIR / f"daily_update_{yyyymmdd(args.date)}_briefing.txt"
    briefing = format_briefing(args.date, results)
    log_path.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    briefing_path.write_text(briefing, encoding="utf-8")
    print(briefing)
    print(json.dumps(results, ensure_ascii=False, indent=2))
    print(f"master={MASTER_FILE}")
    print(f"log={log_path}")
    print(f"briefing={briefing_path}")


if __name__ == "__main__":
    main()
