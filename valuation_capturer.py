import os
import sys
import json
import argparse
from datetime import datetime
from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

def load_credentials():
    credentials_path = "credentials.json"
    if not os.path.exists(credentials_path):
        print(f"Error: '{credentials_path}' file not found.")
        print("Please create a 'credentials.json' file based on the template:")
        print("{\n    \"email\": \"YOUR_EMAIL\",\n    \"password\": \"YOUR_PASSWORD\"\n}")
        sys.exit(1)
        
    with open(credentials_path, "r", encoding="utf-8") as f:
        creds = json.load(f)
        
    email = creds.get("email", "")
    password = creds.get("password", "")
    
    if "YOUR_" in email or not email or not password:
        print("Error: Please update 'credentials.json' with your actual Valley AI credentials.")
        sys.exit(1)
        
    return email, password

def load_stocks():
    stocks_path = "Stocks_Valuation.json"
    if not os.path.exists(stocks_path):
        print(f"Error: '{stocks_path}' file not found.")
        sys.exit(1)
        
    with open(stocks_path, "r", encoding="utf-8") as f:
        return json.load(f)

def navigate_to_stock(page, name, ticker):
    # Strip leading 'A' if it is followed by 6 digits (typical Korean stock code format in FnGuide)
    if ticker.startswith("A") and ticker[1:].isdigit() and len(ticker) == 7:
        ticker = ticker[1:]
        
    candidate_urls = []
    if ":" in ticker:
        candidate_urls.append(f"https://www.valley.town/financials/quote/{ticker}/summary")
    else:
        if ticker.isdigit() and len(ticker) == 6:
            candidate_urls.append(f"https://www.valley.town/financials/quote/{ticker}:KRX/summary")
        else:
            candidate_urls.append(f"https://www.valley.town/financials/quote/{ticker}:NASD/summary")
            candidate_urls.append(f"https://www.valley.town/financials/quote/{ticker}:NYSE/summary")
            
    for url in candidate_urls:
        print(f"Trying direct navigation to: {url}")
        try:
            page.goto(url, wait_until="load")
            page.wait_for_timeout(2000)
            
            # Check for 404
            title = page.title()
            content = page.content()
            if "404" in title or "페이지를 찾을 수 없습니다" in content or "찾을 수 없습니다" in title:
                continue
                
            print(f"Successfully arrived at stock page for {name} ({ticker})")
            return True
        except Exception as e:
            print(f"Error during direct navigation: {e}")
            continue
            
    # Fallback using search input box
    print(f"Direct URL navigation failed for {name} ({ticker}). Trying search fallback...")
    try:
        search_input = page.locator("input[placeholder*='종목명']").first
        if not search_input.is_visible():
            page.goto("https://www.valley.town/lounge", wait_until="load")
            page.wait_for_timeout(2000)
            search_input = page.locator("input[placeholder*='종목명']").first
            
        search_input.click()
        search_input.fill("")
        search_input.type(ticker, delay=100)
        
        # Wait for dropdown item containing ticker
        page.wait_for_selector(f"div[role='listbox'] >> text={ticker}", timeout=5000)
        item = page.locator(f"div[role='listbox'] >> text={ticker}").first
        item.click()
        
        page.wait_for_timeout(3000)
        title = page.title()
        content = page.content()
        if "404" not in title and "페이지를 찾을 수 없습니다" not in content and "찾을 수 없습니다" not in title:
            print(f"Successfully arrived at stock page via search: {page.url}")
            return True
    except Exception as e:
        print(f"Search fallback failed: {e}")
        
    return False

def run_valuation_pipeline(email, password, stocks_dict, metric="PER", headed=False):
    temp_dir = "valley_temp"
    os.makedirs(temp_dir, exist_ok=True)
    
    os.makedirs("컨센서스", exist_ok=True)
    excel_path = os.path.join("컨센서스", "Stocks_Valuation.xlsx")
    
    # 1. Initialize Excel Workbook
    if os.path.exists(excel_path):
        print(f"Loading existing workbook '{excel_path}'...")
        wb = load_workbook(excel_path)
        # Delete old Valuation sheet if it exists
        if "Valuation" in wb.sheetnames:
            print("Removing legacy 'Valuation' sheet...")
            wb.remove(wb["Valuation"])
    else:
        print(f"Creating new workbook '{excel_path}'...")
        wb = Workbook()
        # Remove default sheet to build custom sector sheets
        default_sheet = wb.active
        wb.remove(default_sheet)

    # 2. Setup sheet and column mapping for each sector
    sector_sheets_info = {}
    today_str = datetime.now().strftime("%y.%m.%d")
    
    for sector, stocks in stocks_dict.items():
        if sector in wb.sheetnames:
            ws = wb[sector]
        else:
            ws = wb.create_sheet(title=sector)
            ws["A2"] = "날짜"
            
        # Map stock name to column index (1-based)
        col_map = {}
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=2, column=col).value
            if val:
                col_map[val] = col
                
        # Register new stocks from JSON into headers
        for stock in stocks:
            name = stock["name"]
            if name not in col_map:
                # Find the next empty column in row 2
                next_col = 2
                while ws.cell(row=2, column=next_col).value is not None:
                    next_col += 1
                ws.cell(row=2, column=next_col, value=name)
                col_map[name] = next_col
                
        # Determine target row for today's date in this sheet
        target_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == today_str:
                target_row = r
                break
                
        if not target_row:
            target_row = ws.max_row + 1
            if target_row < 3:
                target_row = 3
            ws.cell(row=target_row, column=1, value=today_str)
            
        print(f"[{sector}] Target row in Excel for date '{today_str}': {target_row}")
        
        # Adjust row height to fit the charts nicely
        ws.row_dimensions[target_row].height = 180  # Height in points
        
        # Save sheet info for playwright processing
        sector_sheets_info[sector] = {
            "target_row": target_row,
            "col_map": col_map
        }

    # 3. Launch Playwright
    with sync_playwright() as p:
        print("Launching browser...")
        browser = p.chromium.launch(headless=not headed)
        context = browser.new_context(viewport={"width": 1440, "height": 900})
        page = context.new_page()
        
        # Log in
        print("Navigating to login page...")
        page.goto("https://www.valley.town/login", wait_until="load")
        
        print("Performing login...")
        page.locator("input[type='email']").fill(email)
        page.locator("input[type='password']").fill(password)
        page.locator("button[type='submit']").click()
        
        # Wait for dashboard/home navigation (wait for URL pathname to change from /login)
        try:
            page.wait_for_function("window.location.pathname !== '/login'", timeout=10000)
            print("Login successful! Redirected to:", page.url)
        except Exception as e:
            print("Warning: Login redirection function timed out. Checking if search bar is visible...")
            page.wait_for_selector("input[placeholder*='종목명']", timeout=5000)
            print("Login validated via search bar presence.")

        # 4. Process each stock by sector
        for sector, stocks in stocks_dict.items():
            ws = wb[sector]
            info = sector_sheets_info[sector]
            target_row = info["target_row"]
            col_map = info["col_map"]
            
            print(f"\n=== Processing Sector: {sector} ===")
            for stock in stocks:
                name = stock["name"]
                ticker = stock["ticker"]
                col_idx = col_map[name]
                
                print(f"\nProcessing stock: {name} ({ticker}) in sheet '{sector}'...")
                
                # Navigate using robust helper
                success = navigate_to_stock(page, name, ticker)
                if not success:
                    print(f"Error: Failed to navigate to stock page for {name} ({ticker}). Skipping.")
                    continue
                
                # Click Valuation tab
                print("Clicking '밸류에이션' tab...")
                valuation_tab = page.locator("text='밸류에이션'").first
                if valuation_tab.is_visible():
                    valuation_tab.click()
                else:
                    page.locator("a:has-text('밸류에이션')").first.click()
                    
                page.wait_for_timeout(3000)  # Wait for charts to load
                
                # Take screenshot of the specified valuation chart card
                print(f"Locating {metric} chart...")
                per_header = page.locator(f"text='{metric}'").first
                if not per_header.is_visible():
                    print(f"Error: Could not locate '{metric}' card. Skipping.")
                    continue
                    
                per_header.scroll_into_view_if_needed()
                page.wait_for_timeout(500)
                
                # Find bounding box of the card container using JS evaluation
                box = per_header.evaluate("""(element) => {
                    let current = element;
                    while (current && current.parentElement) {
                        current = current.parentElement;
                        if (current.innerText.includes('과거 5년') || current.innerText.includes('산업 대비')) {
                            const rect = current.getBoundingClientRect();
                            if (rect.width > 200 && rect.height > 200) {
                                return {
                                    x: rect.x + window.scrollX,
                                    y: rect.y + window.scrollY,
                                    width: rect.width,
                                    height: rect.height
                                };
                            }
                        }
                    }
                    const rect = element.parentElement.getBoundingClientRect();
                    return {
                        x: rect.x + window.scrollX,
                        y: rect.y + window.scrollY,
                        width: rect.width,
                        height: rect.height
                    };
                }""")
                
                if box:
                    # Clip screenshot of the card (ensure metric name is filename safe)
                    safe_metric = metric.replace("/", "_")
                    img_filename = f"{ticker}_{safe_metric}_{today_str.replace('.', '')}.png"
                    img_path = os.path.join(temp_dir, img_filename)
                    print(f"Capturing {metric} chart card screenshot: {img_path}")
                    page.screenshot(path=img_path, clip=box)
                    
                    # Insert image into corresponding column cell of sector sheet
                    col_letter = ws.cell(row=target_row, column=col_idx).column_letter
                    ws.column_dimensions[col_letter].width = 40  # Adjust column width
                    
                    img = OpenpyxlImage(img_path)
                    img.width = 280
                    img.height = 220
                    
                    # Clear cell text first
                    ws.cell(row=target_row, column=col_idx, value="")
                    # Add image
                    ws.add_image(img, f"{col_letter}{target_row}")
                    print(f"Pasted chart image for {name} in sheet '{sector}', cell {col_letter}{target_row}")
                else:
                    print("Error: Failed to calculate card bounding box.")
                    
        browser.close()
        
    # 5. Save the updated workbook
    wb.save(excel_path)
    print(f"\nAll operations completed. Excel updated and saved to: {excel_path}")

def main():
    parser = argparse.ArgumentParser(description="Valley AI 밸류에이션 차트 자동 캡쳐 및 엑셀 삽입 프로그램")
    parser.add_argument("--metric", type=str, default="PER", help="캡쳐할 밸류에이션 지표 (PER, PBR, PSR, P/FCF 등)")
    parser.add_argument("--headed", action="store_true", help="브라우저 화면을 보면서 실행 (디버그용)")
    args = parser.parse_args()
    
    email, password = load_credentials()
    stocks_dict = load_stocks()
    
    run_valuation_pipeline(email, password, stocks_dict, metric=args.metric, headed=args.headed)

if __name__ == "__main__":
    main()
